#!/usr/bin/env python3
"""Table S7 - audited BTLA multiomics evidence across the 43 candidate regulators.

Publication-ready supplementary table built ONLY from the final audit_v2 outputs
and the frozen seed-excluded union of the GREmLN and GENIE3 top-25 rankings. The
evidence is not rerun or reinterpreted here; this module reshapes audited source
records into a compact, redaction-safe table with executable assertions.

Candidate mRNA is contextual (two contrast-specific columns: TCR activation =
TCR vs unstimulated control; incremental BTLA = BTLA+TCR vs TCR). Protein
abundance and phosphosite change (BTLA+TCR vs TCR) are the independent observed
molecular layers. The unavailable BTLA+TCR-vs-unstimulated mRNA contrast is not
a 43-row column; its absence is explained in a footnote.

No restricted values (effect sizes, adjusted p, phosphosite residues / site
identifiers) enter the published table. Phosphosite cells report the number of
distinct qualifying sites and the qualifying timepoints only.
"""
from __future__ import annotations

import re
from pathlib import Path

import pandas as pd

DASH = "\u2014"  # em dash
TP_ORDER = ["2mn", "20mn", "1h", "4h", "24h"]
TP_DISP = {"2mn": "2 min", "20mn": "20 min", "1h": "1 h", "4h": "4 h", "24h": "24 h"}
GROUP_DISP = {"GREmLN_specific": "GREmLN specific", "shared": "Shared",
              "GENIE3_specific": "GENIE3 specific"}
GROUP_ORDER = ["GREmLN_specific", "shared", "GENIE3_specific"]

# expected reconciliation (assertions, not hardcoded replacements)
EXP_MRNA_TCR = {"EGR2", "HIVEP3", "ID2", "JUNB", "PRDM1"}
EXP_PROTEIN = {"SMAP2", "STAT5B"}
EXP_PHOSPHO = {"MYC", "GTF3C2", "HIRIP3", "STAU2", "TFDP1"}
EXP_INDEP = EXP_PROTEIN | EXP_PHOSPHO

RESIDUE_RE = re.compile(r"[A-Za-z0-9]+_[STY]\d+")  # e.g. GENE_S77


def repo_root() -> Path:
    here = Path(__file__).resolve()
    for c in [here.parent, *here.parents]:
        if (c / "scripts").is_dir() and (c / "results").is_dir():
            return c
    return here.parent.parent


def fmt_tps(tps) -> str:
    tps = [t for t in tps if t in TP_DISP]
    tps = sorted(set(tps), key=TP_ORDER.index)
    return ", ".join(TP_DISP[t] for t in tps)


def split_tps(cell) -> list:
    if pd.isna(cell) or not str(cell).strip():
        return []
    return [t.strip() for t in re.split(r"[;,]", str(cell)) if t.strip()]


MEASURED_STATES = {"measured_qualifying", "measured_no_qualifying_signal",
                   "failed_qc", "ambiguous_mapping"}
STATUS_TEXT = {
    "measured_no_qualifying_signal": "Measured, no qualifying change",
    "absent_from_assay_universe": "Absent from assay universe",
    "ambiguous_mapping": "Ambiguous mapping",
    "failed_qc": "Failed quality control",
    "not_measured": "Absent from assay universe",
}


def mrna_cell(status, tp_cell, dir_cell) -> str:
    if status == "measured_qualifying":
        dirs = {d for d in split_tps(dir_cell)}
        word = ("Increased" if dirs == {"up"} else
                "Decreased" if dirs == {"down"} else "Changed")
        return f"{word}: {fmt_tps(split_tps(tp_cell))}"
    return STATUS_TEXT.get(status, "Absent from assay universe")


def indep_cell_protein(status, tp_cell) -> str:
    if status == "measured_qualifying":
        return f"Qualifying association: {fmt_tps(split_tps(tp_cell))}"
    return STATUS_TEXT.get(status, "Absent from assay universe")


def phospho_cell(status, n_sites, tps) -> str:
    if status == "measured_qualifying":
        unit = "site" if n_sites == 1 else "sites"
        return (f"Qualifying association: {fmt_tps(tps)} "
                f"({n_sites} qualifying {unit})")
    return STATUS_TEXT.get(status, "Absent from assay universe")


def main() -> int:
    repo = repo_root()
    out = repo / "results/multiomics/audit_v2"
    rst = out / "local_restricted"

    summ = pd.read_csv(out / "multiomics_candidate_summary_v2.csv")
    mrna = pd.read_csv(out / "mrna_contrast_layers_summary_v2.csv")
    long = pd.read_csv(out / "multiomics_evidence_long_v2.csv")
    union = pd.read_csv(repo / "results/publication_data/top25_union_primary.csv")

    # --- qualifying phospho timepoints (committed long table; residue-free) ---
    ph = long[(long["layer"] == "phosphoproteomics") & (long["qualifying_source"])]
    phospho_tps = ph.groupby("TF")["timepoint"].apply(
        lambda s: sorted(set(s), key=lambda t: TP_ORDER.index(t) if t in TP_ORDER else 99)
    ).to_dict()

    # --- distinct qualifying site COUNT (from restricted FULL; count only) ---
    distinct_sites = {}
    full = rst / "multiomics_candidate_summary_v2.FULL.csv"
    if full.exists():
        f = pd.read_csv(full)
        for _, r in f[f["phospho_status"] == "measured_qualifying"].iterrows():
            toks = split_tps(r.get("phospho_qualifying_site_timepoints"))
            sites = {t.split("@")[0] for t in toks if "@" in t}
            distinct_sites[r["TF"]] = len(sites) if sites else int(r["phospho_qualifying_sites"])
    # fallback to committed (observation) count if FULL unavailable
    for _, r in summ[summ["phospho_status"] == "measured_qualifying"].iterrows():
        distinct_sites.setdefault(r["TF"], int(r["phospho_qualifying_sites"]))

    m = summ.merge(mrna[["TF", "TCR_activation__status", "TCR_activation__qual_timepoints",
                         "TCR_activation__directions", "Incremental_BTLA__status",
                         "Incremental_BTLA__qual_timepoints", "Incremental_BTLA__directions"]],
                   on="TF", how="left")

    rows = []
    for _, r in m.iterrows():
        tf = r["TF"]
        grp = r["candidate_group"]
        in_g = bool(r["in_gremln_top25"])
        in_j = bool(r["in_genie3_top25"])
        g_rank = int(r["gremln_dense_rank"]) if in_g else None
        j_rank = int(r["genie3_dense_rank"]) if in_j else None

        prot_q = r["protein_status"] == "measured_qualifying"
        phos_q = r["phospho_status"] == "measured_qualifying"
        tcr_q = r["TCR_activation__status"] == "measured_qualifying"
        prot_meas = r["protein_status"] in MEASURED_STATES
        phos_meas = r["phospho_status"] in MEASURED_STATES
        mrna_meas = r["TCR_activation__status"] in MEASURED_STATES or \
            r["Incremental_BTLA__status"] in MEASURED_STATES

        # independent molecular association
        if prot_q and phos_q:
            indep = "Yes \u2014 protein and phosphosite"
        elif prot_q:
            indep = "Yes \u2014 protein"
        elif phos_q:
            indep = "Yes \u2014 phosphosite"
        elif prot_meas or phos_meas:
            indep = "No qualifying association"
        else:
            indep = "Not assessable because not measured"

        # interpretation note
        if prot_q and phos_q:
            note = "Independent protein and phosphosite association"
        elif prot_q:
            note = "Independent protein association"
        elif phos_q:
            note = "Independent phosphosite association"
        elif tcr_q:
            note = "Activation-associated mRNA only"
        elif prot_meas or phos_meas or mrna_meas:
            note = "No qualifying signal in measured layers"
        else:
            note = "Sparse assay coverage"

        rows.append({
            "Candidate regulator": tf,
            "Candidate group": GROUP_DISP[grp],
            "GREmLN rank": g_rank if g_rank is not None else DASH,
            "GENIE3 rank": j_rank if j_rank is not None else DASH,
            "TCR activation mRNA": mrna_cell(
                r["TCR_activation__status"], r["TCR_activation__qual_timepoints"],
                r["TCR_activation__directions"]),
            "Incremental BTLA mRNA": mrna_cell(
                r["Incremental_BTLA__status"], r["Incremental_BTLA__qual_timepoints"],
                r["Incremental_BTLA__directions"]),
            "Protein abundance": indep_cell_protein(
                r["protein_status"], r["protein_qualifying_timepoints"]),
            "Phosphosite abundance": phospho_cell(
                r["phospho_status"], distinct_sites.get(tf, 0), phospho_tps.get(tf, [])),
            "Independent molecular association": indep,
            "Interpretation note": note,
            # ordering helpers (dropped before writing)
            "_grp": grp,
            "_g_rank": g_rank if g_rank is not None else 10_000,
            "_j_rank": j_rank if j_rank is not None else 10_000,
        })
    df = pd.DataFrame(rows)

    # --- row order ---
    def order_key(row):
        if row["_grp"] == "GREmLN_specific":
            return (0, row["_g_rank"], row["Candidate regulator"])
        if row["_grp"] == "shared":
            return (1, min(row["_g_rank"], row["_j_rank"]), row["Candidate regulator"])
        return (2, row["_j_rank"], row["Candidate regulator"])
    df["_ord"] = df.apply(order_key, axis=1)
    df = df.sort_values("_ord").reset_index(drop=True)
    grp_series = df["_grp"].copy()
    df = df.drop(columns=["_grp", "_g_rank", "_j_rank", "_ord"])

    # ================= ASSERTIONS =================
    A = []

    def check(name, cond):
        A.append((name, bool(cond)))
        return bool(cond)

    check("exactly_43_rows", len(df) == 43)
    check("each_candidate_once", df["Candidate regulator"].is_unique)
    gc = grp_series.value_counts().to_dict()
    check("group_counts_18_7_18",
          gc.get("GREmLN_specific") == 18 and gc.get("shared") == 7 and
          gc.get("GENIE3_specific") == 18)
    # ranks match canonical union table
    urank = union.set_index("TF")
    rank_ok = True
    for _, r in df.iterrows():
        tf = r["Candidate regulator"]
        gr = r["GREmLN rank"]
        jr = r["GENIE3 rank"]
        if bool(urank.loc[tf, "in_gremln_top25"]):
            rank_ok &= (gr == int(urank.loc[tf, "gremln_dense_rank"]))
        else:
            rank_ok &= (gr == DASH)
        if bool(urank.loc[tf, "in_genie3_top25"]):
            rank_ok &= (jr == int(urank.loc[tf, "genie3_dense_rank"]))
        else:
            rank_ok &= (jr == DASH)
    check("ranks_match_canonical", rank_ok)

    tcr_pos = set(df[df["TCR activation mRNA"].str.startswith(("Increased", "Decreased"))]
                  ["Candidate regulator"])
    check("five_tcr_activation_mrna", tcr_pos == EXP_MRNA_TCR)
    incr_pos = df["Incremental BTLA mRNA"].str.startswith(("Increased", "Decreased")).sum()
    check("zero_incremental_btla_mrna", incr_pos == 0)
    prot_pos = set(df[df["Protein abundance"].str.startswith("Qualifying")]
                   ["Candidate regulator"])
    check("two_protein_qualifying", prot_pos == EXP_PROTEIN)
    phos_pos = set(df[df["Phosphosite abundance"].str.startswith("Qualifying")]
                   ["Candidate regulator"])
    check("five_phospho_qualifying", phos_pos == EXP_PHOSPHO)
    indep_yes = set(df[df["Independent molecular association"].str.startswith("Yes")]
                    ["Candidate regulator"])
    check("seven_independent_association", indep_yes == EXP_INDEP)
    shared_names = set(df[grp_series.values == "shared"]["Candidate regulator"])
    check("no_shared_independent", not (shared_names & indep_yes))
    # not-measured never converted to no-signal, and vice versa: the published cell
    # text must be consistent with the audited source status for every layer.
    msrc = m.set_index("TF")
    nm_ok = True
    for _, r in df.iterrows():
        tf = r["Candidate regulator"]
        pairs = [("TCR activation mRNA", msrc.loc[tf, "TCR_activation__status"]),
                 ("Incremental BTLA mRNA", msrc.loc[tf, "Incremental_BTLA__status"]),
                 ("Protein abundance", msrc.loc[tf, "protein_status"]),
                 ("Phosphosite abundance", msrc.loc[tf, "phospho_status"])]
        for col, st in pairs:
            txt = r[col]
            if st == "absent_from_assay_universe":
                nm_ok &= (txt == "Absent from assay universe")
            if txt == "Measured, no qualifying change":
                nm_ok &= (st == "measured_no_qualifying_signal")
            if txt == "Absent from assay universe":
                nm_ok &= (st in {"absent_from_assay_universe", "not_measured"})
    check("not_measured_preserved", nm_ok)
    # all qualifying cells trace to an audited qualifying_source record
    trace_ok = True
    for tf in tcr_pos:
        trace_ok &= ((long["TF"] == tf) & (long["layer"] == "transcriptomics") &
                     (long["contrast"] == "TCR_vs_UC") & (long["qualifying_source"])).any() or \
                    (mrna.set_index("TF").loc[tf, "TCR_activation__qualifies"])
    for tf in prot_pos:
        trace_ok &= ((long["TF"] == tf) & (long["layer"] == "proteomics") &
                     (long["qualifying_source"])).any()
    for tf in phos_pos:
        trace_ok &= ((long["TF"] == tf) & (long["layer"] == "phosphoproteomics") &
                     (long["qualifying_source"])).any()
    check("qualifying_cells_traceable", trace_ok)
    # no restricted values in the published table
    blob = "\n".join(df.astype(str).apply(lambda s: " ".join(s), axis=1))
    no_residue = not RESIDUE_RE.search(blob)
    no_redacted = "REDACTED" not in blob
    no_forbidden_cols = not any(k in c.lower() for c in df.columns
                                for k in ["log2", "padj", "pvalue", "residue", "site_id"])
    check("no_restricted_values", no_residue and no_redacted and no_forbidden_cols)

    failed = [n for n, ok in A if not ok]
    if failed:
        for n, ok in A:
            print(("PASS" if ok else "FAIL"), n)
        raise SystemExit(f"ASSERTION FAILURE: {failed}")

    # ================= OUTPUTS =================
    stem = "table_S7_audited_multiomics_43_candidates"
    df.to_csv(out / f"{stem}.csv", index=False)

    # markdown (one continuous table)
    md = ["# Table S7. Audited transcriptomic, proteomic and phosphoproteomic "
          "evidence across the 43 candidate regulators", ""]
    md.append("| " + " | ".join(df.columns) + " |")
    md.append("| " + " | ".join(["---"] * len(df.columns)) + " |")
    for _, r in df.iterrows():
        md.append("| " + " | ".join(str(r[c]) for c in df.columns) + " |")
    md.append("")
    md += _footnotes_md()
    (out / f"{stem}.md").write_text("\n".join(md))

    _write_xlsx(df, grp_series.values, out / f"{stem}.xlsx")
    (out / f"{stem}_caption.txt").write_text(_caption(df, tcr_pos, prot_pos, phos_pos, indep_yes))
    (out / f"{stem}_data_dictionary.md").write_text(_data_dictionary())

    # ---- report ----
    print("[Table S7] all assertions passed (", len(A), "checks )")
    print("rows:", len(df), "| groups:", gc)
    print("TCR-activation mRNA responsive (5):", sorted(tcr_pos))
    print("protein qualifying (2):", sorted(prot_pos))
    print("phosphosite qualifying (5):", sorted(phos_pos))
    print("independent molecular association (7):", sorted(indep_yes))
    print("outputs:")
    for ext in ["csv", "xlsx", "md", "_caption.txt", "_data_dictionary.md"]:
        sep = "" if ext.startswith("_") else "."
        print("  ", out / f"{stem}{sep}{ext}")
    return 0


def _footnotes_md():
    return [
        "**Footnotes**", "",
        "a. Candidate mRNA was classified as contextual because the candidates were "
        "nominated as potential upstream regulators of the transcriptional response.",
        "b. Transcriptomic qualifying threshold: adjusted p<0.05 and |log2 fold "
        "change|>log2(1.5).",
        "c. Protein and phosphosite qualifying threshold: adjusted p<0.05 and |log2 "
        "fold change|>log2(1.3).",
        "d. The stricter harmonised sensitivity threshold |log2 fold change|>=0.585 "
        "retained the same seven independent molecular associations.",
        "e. The BTLA+TCR versus unstimulated-control mRNA contrast was not available "
        "from the fitted source models and was not reconstructed from the other "
        "contrasts; it is therefore not shown as a column.",
        "f. \u201cAbsent from assay universe\u201d means the candidate was not measured "
        "and must not be interpreted as evidence of no biological change.",
        "g. Phosphosite cells report the number and timepoints of qualifying sites; "
        "residues and restricted effect estimates remain in the controlled local "
        "audit files.",
        "h. Direction was not interpreted as supportive or opposing because both model "
        "rankings were unsigned.",
    ]


def _caption(df, tcr, prot, phos, indep):
    return (
        "Table S7. Audited transcriptomic, proteomic and phosphoproteomic evidence "
        "across the 43 candidate regulators.\n\n"
        "One row per unique candidate from the frozen seed-excluded union of the "
        "GREmLN and GENIE3 top-25 rankings (GREmLN specific n=18, shared n=7, GENIE3 "
        "specific n=18), grouped and ordered by canonical model rank. Candidate mRNA "
        "is contextual evidence and is shown as two contrast-specific columns: TCR "
        "activation (TCR versus unstimulated control) and incremental BTLA (BTLA+TCR "
        "versus TCR). Protein abundance and phosphosite change (BTLA+TCR versus TCR) "
        "are independent observed molecular layers and define the independent "
        "molecular association call. Missing assay coverage is reported as \u201cAbsent "
        "from assay universe\u201d and is not negative evidence.\n\n"
        f"Of 43 candidates, 41 were measured for mRNA (AHR and MSC absent); five were "
        f"transcriptionally responsive to TCR activation ({', '.join(sorted(tcr))}) and "
        f"none showed a qualifying incremental BTLA mRNA response. Sixteen were "
        f"measured for protein abundance (qualifying: {', '.join(sorted(prot))}) and "
        f"five for phosphosites (qualifying: {', '.join(sorted(phos))}). Seven "
        f"candidates showed an independent molecular association "
        f"({', '.join(sorted(indep))}); none were shared candidates. Because the model "
        f"rankings are unsigned, molecular changes are not labelled supportive or "
        f"opposing. See footnotes a\u2013h for thresholds and definitions."
    )


def _data_dictionary():
    return (
        "# Table S7 data dictionary\n\n"
        "Source: final audit_v2 outputs (multiomics_candidate_summary_v2.csv, "
        "mrna_contrast_layers_summary_v2.csv, multiomics_evidence_long_v2.csv) and the "
        "frozen union results/publication_data/top25_union_primary.csv. No values were "
        "rerun or reinterpreted.\n\n"
        "| Column | Description |\n| --- | --- |\n"
        "| Candidate regulator | Canonical gene symbol. |\n"
        "| Candidate group | GREmLN specific, Shared, or GENIE3 specific. |\n"
        "| GREmLN rank | Canonical GREmLN ordinal (dense) rank if in the GREmLN top 25, "
        "else em dash. |\n"
        "| GENIE3 rank | Canonical GENIE3 ordinal (dense) rank if in the GENIE3 top 25, "
        "else em dash. |\n"
        "| TCR activation mRNA | Contextual. TCR vs unstimulated control. "
        "Increased/Decreased with qualifying timepoint(s), or a controlled status "
        "(Measured, no qualifying change / Absent from assay universe / Ambiguous "
        "mapping / Failed quality control). |\n"
        "| Incremental BTLA mRNA | Contextual. BTLA+TCR vs TCR. Same vocabulary; never "
        "collapsed with the TCR activation contrast. |\n"
        "| Protein abundance | Independent. BTLA+TCR vs TCR. Qualifying association with "
        "timepoint(s) (2 min, 20 min, 4 h, 24 h) or a controlled status. |\n"
        "| Phosphosite abundance | Independent. BTLA+TCR vs TCR. Qualifying association "
        "with timepoint(s) and the number of distinct qualifying sites; residues and "
        "site identifiers are not exposed and sites are not averaged. |\n"
        "| Independent molecular association | Yes \u2014 protein / phosphosite / protein "
        "and phosphosite; No qualifying association; Not assessable because not "
        "measured. True only for a qualifying directly measured protein or phosphosite "
        "change in BTLA+TCR vs TCR. |\n"
        "| Interpretation note | Concise contextualisation (e.g. Activation-associated "
        "mRNA only; Independent protein/phosphosite association; No qualifying signal in "
        "measured layers; Sparse assay coverage). |\n\n"
        "Thresholds: transcriptomics adjusted p<0.05 and |log2FC|>log2(1.5); protein and "
        "phosphosite adjusted p<0.05 and |log2FC|>log2(1.3). Harmonised sensitivity "
        "threshold |log2FC|>=0.585 retained the same seven independent associations. "
        "Restricted effect sizes, adjusted p values and phosphosite residues/site "
        "identifiers remain only in git-ignored local_restricted/ files.\n"
    )


def _write_xlsx(df, groups, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Table S7"
    fills = {"GREmLN_specific": PatternFill("solid", fgColor="EAF1FB"),
             "shared": PatternFill("solid", fgColor="FBF6E7"),
             "GENIE3_specific": PatternFill("solid", fgColor="FBECEC")}
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border
    for i, (_, r) in enumerate(df.iterrows()):
        ws.append([r[c] for c in df.columns])
        rr = i + 2
        for c in range(1, len(df.columns) + 1):
            cell = ws.cell(row=rr, column=c)
            cell.fill = fills[groups[i]]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    widths = [16, 15, 9, 9, 26, 26, 26, 34, 30, 30]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(path)


if __name__ == "__main__":
    raise SystemExit(main())
